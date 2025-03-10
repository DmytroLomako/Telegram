import pandas as pd, os, platform
from .models import User, ResultQuiz, Session
from openpyxl import load_workbook
from openpyxl.styles import Border, Side


def get_test_results(test_id):
    session = Session()
    results = session.query(ResultQuiz).filter_by(test_id=test_id).all()
    session.close()
    test_name = results[0].test_name.split('/')[-1]
    test_names = []
    right_answers = []
    wrong_answers = []
    questions_count = []
    global_results_list = []
    global_result = 0
    session = Session()
    for result in results:
        user = session.query(User).filter_by(id=result.user_id).first()
        test_names.append(user.username)
        right_answers.append(result.right_answers)
        wrong_answers.append(result.wrong_answers)
        questions_count.append(result.right_answers + result.wrong_answers)
    session.close()
    for index in range(len(test_names)):
        result = round(right_answers[index] / (right_answers[index] + wrong_answers[index]) * 100, 1)
        global_results_list.append(result)
    for result in global_results_list:
        global_result += result
    global_result = round(global_result / len(global_results_list), 1)
    df = pd.DataFrame({
        "Ім'я тесту": test_names,
        'Кількість питань': questions_count,
        'Правильні відповіді, кількість': right_answers,
        'Неправильні відповіді, кількість': wrong_answers,
        'Result, %': global_results_list
    })
    
    file_name = f'results_{test_name}.xlsx'
    
    df.to_excel(file_name, index=False)
    
    wb = load_workbook(filename=file_name)
    ws = wb.active
    last_row = ws.max_row + 1
    ws.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=4)
    ws.cell(row=last_row, column=1).value = 'Середній результат, %'
    ws.cell(row=last_row, column=5).value = global_result
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
    wb.save(file_name)
    wb.close()
    
    if platform.system() == 'Windows':
        os.startfile(file_name)
    elif platform.system() == 'Darwin':
        os.system(f'open {file_name}')