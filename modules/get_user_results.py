import pandas as pd, os, platform, ast
from .models import User, Result, Session
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from .read_static import read_json


def get_results(user_id, telegram = False):
    session = Session()
    user = session.query(User).filter_by(id=user_id).first()
    results = session.query(Result).filter_by(user_id=user_id).all()
    session.close()
    username = user.username
    test_names = []
    right_answers = []
    wrong_answers = []
    questions_count = []
    global_results_list = []
    global_result = 0
    for result in results:
        test_names.append(result.test_name.split('/')[-1])
        right_answers.append(result.right_answers)
        wrong_answers.append(result.wrong_answers)
        questions_count.append(result.right_answers + result.wrong_answers)
    for index in range(len(test_names)):
        result = round(right_answers[index] / (right_answers[index] + wrong_answers[index]) * 100, 1)
        global_results_list.append(result)
    for result in global_results_list:
        global_result += result
    global_result = round(global_result / len(global_results_list), 1)
    df = pd.DataFrame({
        "Назва тесту": test_names,
        'Кількість питань': questions_count,
        'Правильні відповіді, кількість': right_answers,
        'Неправильні відповіді, кількість': wrong_answers,
        'Result, %': global_results_list
    })
    
    file_name = f'results_{username}.xlsx'
    
    df.to_excel(file_name, index=False)
    
    wb = load_workbook(filename=file_name)
    ws = wb.active
    last_row = ws.max_row + 1
    ws.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=4)
    ws.cell(row=last_row, column=1).value = 'Середній результат, %'
    ws.cell(row=last_row, column=5).value = global_result
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
    wb.save(file_name)
    wb.close()
    
    if telegram:
        return file_name
    
    if platform.system() == 'Windows':
        os.startfile(file_name)
    elif platform.system() == 'Darwin':
        os.system(f'open {file_name}')
        
def get_one_result(result_id):
    session = Session()
    print(result_id)
    results = session.query(Result).filter_by(id=result_id).first()
    user = session.query(User).filter_by(id=results.user_id).first()
    session.close()
    username = user.username
    test_name = results.test_name.split('/')[-1]
    right_answers = results.right_answers
    wrong_answers = results.wrong_answers
    questions_count = right_answers + wrong_answers
    global_result = round(right_answers / (right_answers + wrong_answers) * 100, 1)
    
    test = read_json(results.test_name)
    questions = test['questions']
    all_variants = []
    all_questions = []
    for question in questions:
        if question['type'] == 'input':
            all_variants.append('')
        else:
            all_variants.append(question['variants'])
        all_questions.append(question['question'])
  
    df = pd.DataFrame({
        'Питання': all_questions,
        'Відповідь': eval(results.result),
        'Варіанти відповідей': all_variants
    })
        
    file_name = f'result_{username}_{test_name}.xlsx'

    df.to_excel(file_name, index=False)

    wb = load_workbook(filename=file_name)
    ws = wb.active
    print(ws.max_column, ws.max_row)

    ws.move_range("A1:C1", rows=1)
    last_row = ws.max_row + 1
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    ws.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=2)
    ws.merge_cells(start_row=last_row + 1, start_column=1, end_row=last_row + 1, end_column=2)
    ws.merge_cells(start_row=last_row + 2, start_column=1, end_row=last_row + 2, end_column=2)
    ws.merge_cells(start_row=last_row + 3, start_column=1, end_row=last_row + 3, end_column=2)
    ws.cell(row=1, column=1).value = test_name
    ws.cell(row=1, column=1).font = ws.cell(row=1, column=1).font.copy(bold=True)
    ws.cell(row=last_row, column=1).value = 'Кількість питань'
    ws.cell(row=last_row, column=1).font = ws.cell(row=1, column=1).font.copy(bold=True)
    ws.cell(row=last_row, column=3).value = questions_count
    ws.cell(row=last_row, column=3).font = ws.cell(row=1, column=1).font.copy(bold=True)
    ws.cell(row=last_row + 1, column=1).value = 'Правильні відповіді, кількість'
    ws.cell(row=last_row + 1, column=1).font = ws.cell(row=1, column=1).font.copy(bold=True)
    ws.cell(row=last_row + 1, column=3).value = right_answers
    ws.cell(row=last_row + 1, column=3).font = ws.cell(row=1, column=1).font.copy(bold=True)
    ws.cell(row=last_row + 2, column=1).value = 'Неправильні відповіді, кількість'
    ws.cell(row=last_row + 2, column=1).font = ws.cell(row=1, column=1).font.copy(bold=True)
    ws.cell(row=last_row + 2, column=3).value = wrong_answers
    ws.cell(row=last_row + 2, column=3).font = ws.cell(row=1, column=1).font.copy(bold=True)
    ws.cell(row=last_row + 3, column=1).value = 'Середній результат, %'
    ws.cell(row=last_row + 3, column=1).font = ws.cell(row=1, column=1).font.copy(bold=True)
    ws.cell(row=last_row + 3, column=3).value = global_result
    ws.cell(row=last_row + 3, column=3).font = ws.cell(row=1, column=1).font.copy(bold=True)
    
    print(ws.max_column, ws.max_row)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for row in ws.iter_rows():
        for cell in row:
            # if cell.value:
            cell.border = border
    wb.save(file_name)
    wb.close()

    return file_name